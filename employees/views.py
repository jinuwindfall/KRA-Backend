import csv
import io
from datetime import date, datetime

from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.db import transaction
from django.utils.dateparse import parse_date
from rest_framework import permissions, status
from rest_framework.authtoken.models import Token
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Department, Employee
from .serializers import DepartmentSerializer, EmployeeCreateSerializer, EmployeeListSerializer


def _employee_payload(emp):
    user = emp.user
    payload = {
        'id': emp.id,
        'emp_id': emp.emp_id,
        'name': user.get_full_name() or user.username,
        'department': emp.department.name if emp.department else 'Unassigned Department',
        'department_id': emp.department.id if emp.department else None,
        'designation': emp.designation,
        'role': emp.role,
        'date_of_joining': str(emp.date_of_joining) if emp.date_of_joining else None,
        'appraiser_name': emp.appraiser.user.get_full_name() if emp.appraiser else '',
        'reviewer_name': emp.reviewer.user.get_full_name() if emp.reviewer else '',
    }
    if emp.role == Employee.ROLE_REVIEWER:
        payload['reviewer_department_ids'] = list(
            emp.reviewer_departments.values_list('id', flat=True)
        )
    return payload


class LoginView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        username = request.data.get('username', '').strip()
        password = request.data.get('password', '')
        if not username or not password:
            return Response({'error': 'Username and password are required.'}, status=status.HTTP_400_BAD_REQUEST)
        user = authenticate(username=username, password=password)
        if not user:
            return Response({'error': 'Invalid credentials.'}, status=status.HTTP_401_UNAUTHORIZED)
        try:
            emp = Employee.objects.select_related(
                'user', 'department', 'appraiser__user', 'reviewer__user'
            ).prefetch_related('reviewer_departments').get(user=user)
        except Employee.DoesNotExist:
            return Response({'error': 'No employee profile linked to this account.'}, status=status.HTTP_400_BAD_REQUEST)
        token, _ = Token.objects.get_or_create(user=user)
        return Response({'token': token.key, 'employee': _employee_payload(emp)})


class ResetPasswordView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        username = request.data.get('username', '').strip()
        emp_id = request.data.get('emp_id', '').strip()
        new_password = request.data.get('new_password', '')

        if not username or not emp_id or not new_password:
            return Response(
                {'error': 'username, emp_id and new_password are required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user = User.objects.get(username=username)
            employee = user.employee
        except (User.DoesNotExist, Employee.DoesNotExist):
            return Response({'error': 'Invalid username or employee ID.'}, status=status.HTTP_400_BAD_REQUEST)

        if employee.emp_id != emp_id:
            return Response({'error': 'Invalid username or employee ID.'}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(new_password)
        user.save(update_fields=['password'])

        Token.objects.filter(user=user).delete()
        return Response({'message': 'Password reset successful. Please sign in with your new password.'})


class ChangePasswordView(APIView):
    """Authenticated user changes their own password."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        current_password = request.data.get('current_password', '')
        new_password = request.data.get('new_password', '')

        if not current_password or not new_password:
            return Response(
                {'error': 'current_password and new_password are required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if len(new_password) < 4:
            return Response(
                {'error': 'New password must be at least 4 characters.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = request.user
        if not user.check_password(current_password):
            return Response({'error': 'Current password is incorrect.'}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(new_password)
        user.save(update_fields=['password'])
        # Invalidate all existing tokens so the user must log in again
        Token.objects.filter(user=user).delete()
        return Response({'message': 'Password changed successfully. Please log in with your new password.'})


class MeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            emp = Employee.objects.select_related(
                'user', 'department', 'appraiser__user', 'reviewer__user'
            ).prefetch_related('reviewer_departments').get(user=request.user)
        except Employee.DoesNotExist:
            return Response({'error': 'No employee profile.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(_employee_payload(emp))


class DepartmentListCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        departments = Department.objects.all()
        serializer = DepartmentSerializer(departments, many=True)
        return Response(serializer.data)

    def post(self, request):
        try:
            emp = request.user.employee
        except Employee.DoesNotExist:
            return Response({'error': 'No employee profile.'}, status=status.HTTP_403_FORBIDDEN)
        if emp.role not in {Employee.ROLE_REVIEWER, Employee.ROLE_HR}:
            return Response({'error': 'Only reviewers or HR can create departments.'}, status=status.HTTP_403_FORBIDDEN)
        serializer = DepartmentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class DepartmentDetailView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        try:
            emp = request.user.employee
        except Employee.DoesNotExist:
            return Response({'error': 'No employee profile.'}, status=status.HTTP_403_FORBIDDEN)

        if emp.role != Employee.ROLE_HR:
            return Response({'error': 'Only HR can edit departments.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            department = Department.objects.get(pk=pk)
        except Department.DoesNotExist:
            return Response({'error': 'Department not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = DepartmentSerializer(department, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        try:
            emp = request.user.employee
        except Employee.DoesNotExist:
            return Response({'error': 'No employee profile.'}, status=status.HTTP_403_FORBIDDEN)

        if emp.role != Employee.ROLE_HR:
            return Response({'error': 'Only HR can remove departments.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            department = Department.objects.get(pk=pk)
        except Department.DoesNotExist:
            return Response({'error': 'Department not found.'}, status=status.HTTP_404_NOT_FOUND)

        department.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class EmployeeListCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            emp = request.user.employee
        except Employee.DoesNotExist:
            return Response({'error': 'No employee profile.'}, status=status.HTTP_403_FORBIDDEN)
        if emp.role == Employee.ROLE_APPRAISER:
            dept_ids = list(emp.appraiser_departments.values_list('id', flat=True))
            employees = Employee.objects.filter(
                appraiser=emp
            )
            if dept_ids:
                employees = employees.filter(department_id__in=dept_ids)
            employees = employees.select_related(
                'user', 'department', 'appraiser__user', 'reviewer__user'
            ).prefetch_related(
                'reviewer_departments', 'appraiser_departments'
            ).order_by('department__name', 'user__first_name', 'user__last_name')
        elif emp.role == Employee.ROLE_REVIEWER:
            dept_ids = list(emp.reviewer_departments.values_list('id', flat=True))
            reviewer_qs = Employee.objects.filter(reviewer=emp)
            if dept_ids:
                reviewer_qs = reviewer_qs.filter(department_id__in=dept_ids)
            employees = reviewer_qs.select_related(
                'user', 'department', 'appraiser__user', 'reviewer__user'
            ).prefetch_related(
                'reviewer_departments', 'appraiser_departments'
            ).order_by('department__name', 'user__first_name', 'user__last_name')
        elif emp.role == Employee.ROLE_HR:
            employees = Employee.objects.all().select_related(
                'user', 'department', 'appraiser__user', 'reviewer__user'
            ).prefetch_related(
                'reviewer_departments', 'appraiser_departments'
            ).order_by('department__name', 'user__first_name', 'user__last_name')
        else:
            return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        serializer = EmployeeListSerializer(employees, many=True)
        return Response(serializer.data)

    def post(self, request):
        try:
            emp = request.user.employee
        except Employee.DoesNotExist:
            return Response({'error': 'No employee profile.'}, status=status.HTTP_403_FORBIDDEN)
        if emp.role not in {Employee.ROLE_REVIEWER, Employee.ROLE_HR}:
            return Response({'error': 'Only reviewers or HR can create employees.'}, status=status.HTTP_403_FORBIDDEN)
        serializer = EmployeeCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        new_emp = serializer.save()
        new_emp = Employee.objects.select_related(
            'user', 'department', 'appraiser__user', 'reviewer__user'
        ).prefetch_related('reviewer_departments', 'appraiser_departments').get(pk=new_emp.pk)
        return Response(EmployeeListSerializer(new_emp).data, status=status.HTTP_201_CREATED)


class EmployeeDetailView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _get_employee(self, request, pk):
        try:
            caller = request.user.employee
        except Employee.DoesNotExist:
            return None, Response({'error': 'No employee profile.'}, status=status.HTTP_403_FORBIDDEN)

        if caller.role == Employee.ROLE_HR:
            employee = Employee.objects.filter(pk=pk).select_related(
                'user', 'department', 'appraiser__user', 'reviewer__user'
            ).prefetch_related('reviewer_departments', 'appraiser_departments').first()
        elif caller.role == Employee.ROLE_REVIEWER:
            dept_ids = list(caller.reviewer_departments.values_list('id', flat=True))
            reviewer_qs = Employee.objects.filter(pk=pk, reviewer=caller)
            if dept_ids:
                reviewer_qs = reviewer_qs.filter(department_id__in=dept_ids)
            employee = reviewer_qs.select_related(
                'user', 'department', 'appraiser__user', 'reviewer__user'
            ).prefetch_related('reviewer_departments', 'appraiser_departments').first()
        else:
            employee = None

        if not employee:
            return None, Response({'error': 'Employee not found or permission denied.'}, status=status.HTTP_404_NOT_FOUND)

        return employee, None

    def patch(self, request, pk):
        employee, error_response = self._get_employee(request, pk)
        if error_response:
            return error_response

        serializer = EmployeeCreateSerializer(employee, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        updated_employee = serializer.save()
        return Response(EmployeeListSerializer(updated_employee).data)

    def delete(self, request, pk):
        employee, error_response = self._get_employee(request, pk)
        if error_response:
            return error_response

        if employee.user_id == request.user.id:
            return Response({'error': 'You cannot delete your own account.'}, status=status.HTTP_400_BAD_REQUEST)

        user = employee.user
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class EmployeeBulkImportView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    REQUIRED_HEADERS = {'emp_id', 'name', 'designation', 'role'}

    # Maps normalized human-readable column headers → internal field names
    HEADER_MAP = {
        'emp id': 'emp_id',
        'name of the employee': 'name',
        'email id': 'email',
        'desigination': 'designation',
        'designation': 'designation',
        'date of joining': 'date_of_joining',
        'appraiser': 'appraiser_name',
        'reviewer': 'reviewer_name',
    }

    @staticmethod
    def _normalize_header(header):
        raw = str(header or '').strip().lower()
        return EmployeeBulkImportView.HEADER_MAP.get(raw, raw)

    @staticmethod
    def _normalize_role(value):
        return str(value or '').strip().lower()

    # Role priority: highest wins when multiple roles are listed
    ROLE_PRIORITY = {
        Employee.ROLE_HR: 4,
        Employee.ROLE_REVIEWER: 3,
        Employee.ROLE_APPRAISER: 2,
        Employee.ROLE_STAFF: 1,
    }

    def _resolve_role(self, raw_value):
        """
        Parse a potentially comma-separated role string like 'Appraiser, Reviewer'
        and return (primary_role, [all_roles]) using priority order.
        """
        parts = [p.strip().lower() for p in str(raw_value or '').split(',') if p.strip()]
        valid = [r for r in parts if r in self.ROLE_PRIORITY]
        if not valid:
            return None, parts
        primary = max(valid, key=lambda r: self.ROLE_PRIORITY[r])
        return primary, valid

    @staticmethod
    def _parse_bool(value, default=True):
        if value in (None, ''):
            return default
        if isinstance(value, bool):
            return value
        str_val = str(value).strip().lower()
        if str_val in {'1', 'true', 'yes', 'y'}:
            return True
        if str_val in {'0', 'false', 'no', 'n'}:
            return False
        return default

    @staticmethod
    def _parse_date_value(value):
        if value in (None, ''):
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        str_val = str(value).strip()
        # Try Django's parse_date (YYYY-MM-DD) first
        parsed = parse_date(str_val)
        if parsed:
            return parsed.isoformat()
        # Try DD-Month-YYYY and DD/Month/YYYY  e.g. 01-December-2023
        for fmt in ('%d-%B-%Y', '%d/%B/%Y', '%d-%b-%Y', '%d/%b/%Y',
                    '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(str_val, fmt).date().isoformat()
            except ValueError:
                continue
        return None

    def _rows_from_csv(self, upload):
        content = upload.read().decode('utf-8-sig')
        stream = io.StringIO(content)
        reader = csv.DictReader(stream)
        headers = {self._normalize_header(h) for h in (reader.fieldnames or [])}
        missing = self.REQUIRED_HEADERS - headers
        if missing:
            raise ValueError(f"Missing required column(s): {', '.join(sorted(missing))}")
        rows = []
        for row in reader:
            normalized = {self._normalize_header(k): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
            rows.append(normalized)
        return rows

    def _rows_from_xlsx(self, upload):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError('XLSX import requires openpyxl. Install it in backend environment.') from exc

        workbook = load_workbook(upload, data_only=True)
        sheet = workbook.active
        raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [self._normalize_header(h) for h in raw_headers]
        missing = self.REQUIRED_HEADERS - set(headers)
        if missing:
            raise ValueError(f"Missing required column(s): {', '.join(sorted(missing))}")

        rows = []
        for excel_row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell in (None, '') for cell in excel_row):
                continue
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                row[header] = excel_row[index] if index < len(excel_row) else None
            rows.append(row)
        return rows

    def _get_rows(self, upload):
        filename = (upload.name or '').lower()
        if filename.endswith('.csv'):
            return self._rows_from_csv(upload)
        if filename.endswith('.xlsx'):
            return self._rows_from_xlsx(upload)
        raise ValueError('Unsupported file type. Please upload a .csv or .xlsx file.')

    @transaction.atomic
    def post(self, request):
        try:
            caller = request.user.employee
        except Employee.DoesNotExist:
            return Response({'error': 'No employee profile.'}, status=status.HTTP_403_FORBIDDEN)
        if caller.role != Employee.ROLE_HR:
            return Response({'error': 'Only HR can import employees.'}, status=status.HTTP_403_FORBIDDEN)

        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': 'file is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = self._get_rows(upload)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if not rows:
            return Response({'error': 'The uploaded sheet is empty.'}, status=status.HTTP_400_BAD_REQUEST)

        created = []
        relation_updates = []
        role_warnings = []
        valid_roles = {choice[0] for choice in Employee.ROLE_CHOICES}

        for index, row in enumerate(rows, start=2):
            emp_id = str(row.get('emp_id') or '').strip()
            name = str(row.get('name') or '').strip()
            designation = str(row.get('designation') or '').strip()

            role, all_roles = self._resolve_role(row.get('role'))

            if not emp_id or not name or not designation or not role:
                raw_role = str(row.get('role') or '').strip()
                return Response(
                    {'error': f'Row {index}: emp_id, name, designation and role are required. '
                              f'Allowed roles: {", ".join(sorted(valid_roles))}.'
                              + (f' Got role: "{raw_role}"' if raw_role else '')},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if len(all_roles) > 1:
                role_warnings.append(
                    f"Row {index} ({emp_id}): multiple roles {all_roles} — assigned as '{role}' (highest priority)."
                )

            department = None
            department_name = str(row.get('department') or '').strip()
            if department_name:
                department, _ = Department.objects.get_or_create(name=department_name)

            payload = {
                'emp_id': emp_id,
                'name': name,
                'email': str(row.get('email') or '').strip(),
                'designation': designation,
                'role': role,
                'department': department.id if department else None,
                'date_of_joining': self._parse_date_value(row.get('date_of_joining')),
                'is_active': self._parse_bool(row.get('is_active'), default=True),
                'gender': str(row.get('gender') or '').strip() or None,
            }

            # Skip rows where emp_id already exists — but still update appraiser/reviewer
            if Employee.objects.filter(emp_id=emp_id).exists():
                role_warnings.append(f"Row {index} ({emp_id}): already exists — skipped creation.")
                existing_emp = Employee.objects.get(emp_id=emp_id)
                relation_updates.append({
                    'employee': existing_emp,
                    'appraiser_name': str(row.get('appraiser_name') or '').strip(),
                    'reviewer_name': str(row.get('reviewer_name') or '').strip(),
                })
                continue

            serializer = EmployeeCreateSerializer(data=payload)
            if not serializer.is_valid():
                return Response(
                    {'error': f'Row {index}: {serializer.errors}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            employee = serializer.save()
            created.append(employee)
            relation_updates.append({
                'employee': employee,
                'appraiser_name': str(row.get('appraiser_name') or '').strip(),
                'reviewer_name': str(row.get('reviewer_name') or '').strip(),
            })

        warnings = []
        for rel in relation_updates:
            employee = rel['employee']
            appraiser_name = rel['appraiser_name']
            reviewer_name = rel['reviewer_name']

            if appraiser_name:
                # Look up by full name (case-insensitive) — try first+last split, then username
                name_parts = appraiser_name.strip().split(maxsplit=1)
                first = name_parts[0]
                last = name_parts[1] if len(name_parts) > 1 else ''
                appraiser = Employee.objects.filter(
                    user__first_name__iexact=first,
                    user__last_name__iexact=last,
                ).first()
                if not appraiser:
                    appraiser = Employee.objects.filter(
                        user__username__iexact=appraiser_name.strip()
                    ).first()
                if appraiser:
                    employee.appraiser = appraiser
                else:
                    warnings.append(f"{employee.emp_id}: appraiser '{appraiser_name}' not found.")

            if reviewer_name:
                name_parts = reviewer_name.strip().split(maxsplit=1)
                first = name_parts[0]
                last = name_parts[1] if len(name_parts) > 1 else ''
                reviewer = Employee.objects.filter(
                    user__first_name__iexact=first,
                    user__last_name__iexact=last,
                ).first()
                if not reviewer:
                    reviewer = Employee.objects.filter(
                        user__username__iexact=reviewer_name.strip()
                    ).first()
                if reviewer:
                    employee.reviewer = reviewer
                else:
                    warnings.append(f"{employee.emp_id}: reviewer '{reviewer_name}' not found.")

            if appraiser_name or reviewer_name:
                employee.save(update_fields=['appraiser', 'reviewer'])

        return Response({
            'created_count': len(created),
            'created': EmployeeListSerializer(created, many=True).data,
            'credentials_rule': 'username is generated from name, password is employee ID.',
            'warnings': role_warnings + warnings,
        }, status=status.HTTP_201_CREATED)


class ReviewerDepartmentsView(APIView):
    """Set the departments a reviewer can access. HR-only."""
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        try:
            caller = request.user.employee
        except Employee.DoesNotExist:
            return Response({'error': 'No employee profile.'}, status=status.HTTP_403_FORBIDDEN)
        if caller.role != Employee.ROLE_HR:
            return Response({'error': 'Only HR can assign reviewer departments.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            reviewer = Employee.objects.get(pk=pk, role=Employee.ROLE_REVIEWER)
        except Employee.DoesNotExist:
            return Response({'error': 'Reviewer not found.'}, status=status.HTTP_404_NOT_FOUND)
        dept_ids = request.data.get('department_ids', [])
        departments = Department.objects.filter(id__in=dept_ids)
        reviewer.reviewer_departments.set(departments)
        return Response({
            'id': reviewer.id,
            'reviewer_department_ids': list(reviewer.reviewer_departments.values_list('id', flat=True)),
        })


class AppraiserDepartmentsView(APIView):
    """Set the departments an appraiser handles. HR-only."""
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        try:
            caller = request.user.employee
        except Employee.DoesNotExist:
            return Response({'error': 'No employee profile.'}, status=status.HTTP_403_FORBIDDEN)
        if caller.role != Employee.ROLE_HR:
            return Response({'error': 'Only HR can assign appraiser departments.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            appraiser = Employee.objects.get(pk=pk, role=Employee.ROLE_APPRAISER)
        except Employee.DoesNotExist:
            return Response({'error': 'Appraiser not found.'}, status=status.HTTP_404_NOT_FOUND)
        dept_ids = request.data.get('department_ids', [])
        departments = Department.objects.filter(id__in=dept_ids)
        appraiser.appraiser_departments.set(departments)
        return Response({
            'id': appraiser.id,
            'appraiser_department_ids': list(appraiser.appraiser_departments.values_list('id', flat=True)),
        })


class DepartmentManagersView(APIView):
    """Given a department_id, return the appraiser and reviewer assigned to it."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        dept_id = request.query_params.get('department_id')
        if not dept_id:
            return Response({'error': 'department_id is required.'}, status=status.HTTP_400_BAD_REQUEST)
        appraiser = Employee.objects.filter(
            role=Employee.ROLE_APPRAISER,
            appraiser_departments__id=dept_id,
        ).select_related('user').first()
        reviewer = Employee.objects.filter(
            role=Employee.ROLE_REVIEWER,
            reviewer_departments__id=dept_id,
        ).select_related('user').first()
        return Response({
            'appraiser_id': appraiser.id if appraiser else None,
            'appraiser_name': appraiser.user.get_full_name() if appraiser else '',
            'reviewer_id': reviewer.id if reviewer else None,
            'reviewer_name': reviewer.user.get_full_name() if reviewer else '',
        })


